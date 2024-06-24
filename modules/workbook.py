from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles import numbers
import openpyxl


class WorkbookManager:

    @staticmethod
    def create_workbook():
        return openpyxl.Workbook()

    @staticmethod
    def load_workbook(filename):
        try:
            return openpyxl.load_workbook(filename)
        except Exception as e:
            raise Exception('Ошибка при загрузке книги') from e

    @staticmethod
    def create_sheet(workbook, name):
        return workbook.create_sheet(name)

    @staticmethod
    def format_sheet(sheet, type=None):
        try:
            sheet.auto_filter.ref = sheet.dimensions
            sheet.freeze_panes = 'A2'

            for cell in sheet[1]:
                cell.fill = PatternFill(
                    start_color='5F4876', end_color='5F4876', fill_type='solid')
                cell.font = Font(color='FFFFFF')
                cell.alignment = Alignment(horizontal='left')

            percent_format = numbers.FORMAT_PERCENTAGE_00

            if type == 'affinity':
                for cell in sheet['E']:
                    cell.number_format = percent_format

        except Exception as e:
            raise Exception('Ошибка при форматировании рабочего листа') from e

    @staticmethod
    def save_workbook(workbook, name):
        try:
            del workbook['Sheet']
        except:
            pass

        workbook.save(name)
