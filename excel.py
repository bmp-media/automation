import openpyxl
from modules.workbook import WorkbookManager
from modules.paths import PathManager

wb1 = WorkbookManager.load_workbook(PathManager.open_file_dialog_solo())
wb2 = WorkbookManager.load_workbook(PathManager.open_file_dialog_solo())


# Проходим по всем листам в первом файле
for sheet_name in wb1.sheetnames:
    sheet1 = wb1[sheet_name]
    sheet2 = wb2[sheet_name]

    # Проходим по всем ячейкам в листе
    for row in range(1, sheet1.max_row + 1):
        for col in range(1, sheet1.max_column + 1):
            cell1 = sheet1.cell(row=row, column=col)
            cell2 = sheet2.cell(row=row, column=col)

            # Сравниваем значения ячеек
            if cell1.value != cell2.value:
                # Если значения отличаются, закрашиваем ячейку в красный цвет
                cell1.fill = openpyxl.styles.PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

# Сохраняем Excel книгу
file_path_to_save = PathManager.save_file_dialog()

if file_path_to_save:
    WorkbookManager.save_workbook(wb1, file_path_to_save)
    print(f'Файл успешно сохранен: {file_path_to_save}')
else:
    print('Файл не сохранен.')
